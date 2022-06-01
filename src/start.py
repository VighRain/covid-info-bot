"""Основной скрипт для работы бота"""
import telebot
from bot_token import TOKEN
from datetime import datetime, timedelta
from openpyxl import load_workbook

bot = telebot.TeleBot(TOKEN)

def sql_init():
        cursor = connection.cursor()
        

@bot.message_handler(commands=['start'])
def start(message):
        """
        Вывод на экран начального меню бота.

        Выводится приветствие из /data/start.txt
        и отображаются 5 кнопок для взаимодействия с ботом.
        
        """
        with open('../data/start.txt', encoding="utf-8") as f:
                start = f.read()
        markup = telebot.types.InlineKeyboardMarkup()
        info_button = markup.add(telebot.types.InlineKeyboardButton(text='Info', callback_data=1))
        stats_button = markup.add(telebot.types.InlineKeyboardButton(text='Statistics', callback_data=2))
        health_button = markup.add(telebot.types.InlineKeyboardButton(text='Health Services', callback_data=4))
        bot.send_message(message.chat.id, text=start, reply_markup=markup)
        
@bot.message_handler(commands=['info'])
def info(message):
        """
        Ввывод общей информации о вирусе и эпидемии.
        
        Выводится из /data/info.txt
        
        """
        with open('../data/info.txt', encoding="utf-8") as f:
                info = f.read()
        bot.send_message(message.chat.id, text=info)

@bot.message_handler(commands=['stats'])
def stats(message):
        """
        Показ статистики по выбранной стране.

        """

        sent_msg = bot.send_message(message.chat.id, "Which country to view?")
        bot.register_next_step_handler(sent_msg, country_handler)

def country_stats)handler(message):
        
        country = message.text

        covid_set = '../data/owid-covid-data.xlsx'
        covid_book = load_workbook(covid_set, read_only=True)
        sheet = covid_book.active
        
        date = datetime.now()-timedelta(days=1)
        date = date.strftime("%Y-%m-%d")
        print(date)
        print(country)
        
        for row in sheet.iter_rows(min_col=2, min_row=2):
                if row[1].value == country:
                        if row[2].value == date:
                                print(f"Total cases today: {row[3].value}")
                                bot.send_message(message.chat.id, f"Total cases today: {row[3].value}")
                        else:
                                print("n/a")
                                break


@bot.message_handler(commands=['health'])
def health(message):

        sent_msg = bot.send_message(message.chat.id, "Where are you right now?")
        bot.register_next_step_handler(sent_msg, country_handler)

def country_stats_handler(message):
        
        country = message.text

        covid_set = '../data/owid-covid-data.xlsx'
        covid_book = load_workbook(covid_set, read_only=True)
        sheet = covid_book.active
        
        date = datetime.now()-timedelta(days=1)
        date = date.strftime("%Y-%m-%d")
        print(date)
        print(country)
        
        for row in sheet.iter_rows(min_col=2, min_row=2):
                if row[1].value == country:
                        if row[2].value == date:
                                print(f"Total cases today: {row[3].value}")
                                bot.send_message(message.chat.id, f"Total cases today: {row[3].value}")
                        else:
                                print("n/a")
                                break

@bot.callback_query_handler(func=lambda call: True)
def query_handler(call):
        """
        Обработчик меню.

        Получает callback_data из нажатой кнопки
        и выполняет соответствующее кнопки действие.
        
        """

        bot.answer_callback_query(callback_query_id=call.id)
        
        if call.data == '1':
                info(call.message)
        elif call.data == '2':
                stats(call.message)
        elif call.data == '3':
                tourism(call.message)
        elif call.data == '4':
                health(call.message)


if __name__ == '__main__':
        bot.infinity_polling()
